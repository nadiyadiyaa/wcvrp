from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from master.models import *
from decimal import Decimal


class Command(BaseCommand):
    help = "Command to initialize truck"

    def handle(self, *args, **kwargs):
        path = "master/templates/excel/truk_dan_kapasitas.xlsx"
        wb = load_workbook(path, data_only=True)

        sheet = wb['Truk dan Kapasitas']
        num = 0
        for key, *values in sheet.iter_rows():
            if num > 0:
                kec = Kecamatan.objects.filter(name=values[0].value).first()
                type = TruckType.objects.filter(
                    truck_type=values[3].value).first()

                truck = Truck(**{
                    'kecamatan_id': kec.id,
                    'no_pol': values[1].value,
                    'year': values[2].value,
                    'type_id': type.id,
                    # 'capacity': Decimal(str(values[4].value).replace('m³', '').strip()),
                    'truck_class_id': values[6].value,
                    'created_by': 'System',
                    'modified_by': 'System',
                })
                truck.save()

            num += 1
