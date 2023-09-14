from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from master.models import *
from decimal import Decimal


class Command(BaseCommand):
    help = "Command to initialize other places"

    def handle(self, *args, **kwargs):
        path = "master/templates/excel/koordinat_tps_2.xlsx"
        wb = load_workbook(path, data_only=True)

        sheet = wb['Koordinat TPS, TPA, Depot']
        num = 0
        for key, *values in sheet.iter_rows():
            if num > 0 and num < 67:
                place = Place(**{
                    'nodes': int(key.value),
                    'name': values[0].value,
                    'location': values[1].value,
                    'latitude': values[2].value,
                    'longitude': values[3].value,
                    'volume': float(values[4].value),
                    'created_by': 'System',
                    'modified_by': 'System',
                })
                place.save()

            num += 1
