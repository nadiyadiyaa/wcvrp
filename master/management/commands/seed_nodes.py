from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from master.models import *
from decimal import Decimal


class Command(BaseCommand):
    help = "Command to initialize other places"

    def handle(self, *args, **kwargs):
        path = "master/templates/excel/jarak_nodes.xlsx"
        wb = load_workbook(path)

        sheet = wb['Jarak antar Titik Depot-TPA-TPS']

        num = 0
        for key, *values in sheet.iter_rows():
            if num > 0 and num < 69:
                from_node = key.value
                fp = Place.objects.filter(nodes=from_node).first()

                for idx in range(0, num - 1):
                    value = values[idx].value
                    if value:
                        tp = Place.objects.filter(nodes=idx).first()
                        distance = PlaceDistance(
                            from_place_id=fp.id,
                            to_place_id=tp.id,
                            distance=Decimal(values[idx].value),
                            created_by='System',
                            modified_by='System',
                        )
                        distance.save()

            num += 1
